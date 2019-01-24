#importamos la libreria
import openpyxl
import MySQLdb
import time
from datetime import datetime, date, time, timedelta
#caragamos la hoja de calculo en memoria
doc = openpyxl.load_workbook('PAGOS.xlsx')

#coneccion al db
db = MySQLdb.connect(host="172.18.55.99",    # tu host, usualmente localhost
                     user="comandato",         # tu usuario
                     passwd="comandato123",  # tu password
                     db="SISTEMECUADOR_ATM")        # el nombre de la base de datos

cur = db.cursor()
cur1 = db.cursor()
cur2 = db.cursor()

#escogemos una hoja para trabajar con su respectivo nombre de variable
hoja = doc.get_sheet_by_name('REVISAR')
hoja2 = doc.get_sheet_by_name('RESULTADO')

a=0
def days_between(d1, d2):
    return abs(d2 - d1).days

def rev_segmento(altura,segm):
    segmento=''
    if segm=='No':
        if altura=='1. 0 - 30 días' or altura=='2. 31 - 60 días' or altura=='3. 61 - 90 días' or altura=='4. 91 - 180 días':
            segmento='DE 1 A 180 DIAS'
        elif altura=='5. 181 - 270 días' or altura=='6. 271 - 360 días' or altura=='7. 361 - 720 días' or altura=='8. Mayor a 720 días':
            segmento='MAYOR A 180 DIAS'
    elif segm=='Si':
        segmento='MENOR CUANTIA'
    else:
        segmento='NO IDENTIFICADO'

    return segmento

def consulta_ivr(fech_ini,fech_fin,cedula):
    consultar_IVR_gestion="SELECT FECHA_GESTION FROM SISTEMECUADOR_ATM.GESTION where  FECHA_GESTION between '"+str(fech_ini)+"' and '"+str(fech_fin)+"' and NRO_IDENTIFICACION_CLIENTE='"+cedula+"' and TIEMPO_GESTION>0 and (RESPUESTA_OBTENIDA='IVR' OR RESPUESTA_OBTENIDA='CONTESTO IVR') order by FECHA_GESTION limit 1;"
    return consultar_IVR_gestion

def consultar_gest(fech_ini,fech_fin,cedula):
    calltype="'ACUERDO DE CONVENIO','LOCALIZADO SIN ACUERDO','MENSAJE CON TERCERO','YA PAGO','ACUERDO DE PAGO','RECLAMO VENDIO VEHICULO','RECLAMO'"
    consultar_gestion="SELECT RESPUESTA_OBTENIDA FROM SISTEMECUADOR_ATM.GESTION where  FECHA_GESTION between '"+str(fech_ini)+"' and '"+str(fech_fin)+"' and NRO_IDENTIFICACION_CLIENTE='"+cedula+"' and RESPUESTA_OBTENIDA in ("+calltype+") order by FECHA_GESTION desc limit 1;"
    return consultar_gestion

for i in range(1,hoja.max_row):

    cedula=str(hoja.cell(row=i+1,column=2).value)
    diasMora=int(hoja.cell(row=i+1,column=6).value)
    altura=str(hoja.cell(row=i+1,column=7).value)
    menorCuantia=str(hoja.cell(row=i+1,column=9).value)
    seg=rev_segmento(altura,menorCuantia)
    fech_asig=hoja.cell(row=i+1,column=4).value
    fech_pago=hoja.cell(row=i+1,column=5).value
    valor=hoja.cell(row=i+1,column=8).value
    val_hon=0
    fec_ivr='00/00/00'
    fech_gst='00/00/00'
    nec_gst='si'

    contar_gestion="SELECT COUNT(*) FROM SISTEMECUADOR_ATM.GESTION where  NRO_IDENTIFICACION_CLIENTE='"+cedula+"' and (RESPUESTA_OBTENIDA!='IVR' OR RESPUESTA_OBTENIDA!='CONTESTO IVR') AND FECHA_GESTION between '"+str(fech_asig)+"' and '"+str(fech_pago)+"';"

    if seg=='MENOR CUANTIA':
        aplData=str(hoja.cell(row=i+1,column=10).value)
        dias_entre=days_between(fech_asig.date(),fech_pago.date())
        if(dias_entre<8):
            querry_ivr=consulta_ivr(fech_asig,str(fech_pago.date()+timedelta(days=1)),cedula)
        else:
            querry_ivr=consulta_ivr(str(fech_pago.date()-timedelta(days=7)),str(fech_pago.date()+timedelta(days=1)),cedula)
        querry_gest=consultar_gest(str(fech_pago.date()- timedelta(days=30)),str(fech_pago.date()+timedelta(days=1)),cedula)
        val_hon=2
        if diasMora>90:
            nec_gst='no'
    elif seg=='DE 1 A 180 DIAS':
        aplData=str(hoja.cell(row=i+1,column=14).value)
        dias_entre=days_between(fech_asig.date(),fech_pago.date())
        if(dias_entre<8):
            querry_ivr=consulta_ivr(str(fech_asig),str(fech_pago.date()+timedelta(days=1)),cedula)
        else:
            querry_ivr="SELECT '0' "
        querry_gest=consultar_gest(str(fech_pago.date()- timedelta(days=15)),str(fech_pago.date()+timedelta(days=1)),cedula)
        val_hon=valor*0.037
    elif seg=='MAYOR A 180 DIAS':
        aplData=str(hoja.cell(row=i+1,column=14).value)
        dias_entre=days_between(fech_asig.date(),fech_pago.date())
        if(dias_entre<8):
            querry_ivr=consulta_ivr(str(fech_asig),str(fech_pago.date()+timedelta(days=1)),cedula)
        else:
            querry_ivr=consulta_ivr(str(fech_pago.date()-timedelta(days=7)),str(fech_pago.date()+timedelta(days=1)),cedula)
        querry_gest=consultar_gest(str(fech_pago.date()- timedelta(days=30)),str(fech_pago.date()+timedelta(days=1)),cedula)
        val_hon=valor*0.045
    else:
        aplData='NO INDENTIFICADO'

    try:
        fech_ivr=cur.execute(querry_ivr)
        db.commit()
        registro = cur.fetchone()
        while (registro != None):
            fec_ivr=registro[0]
            registro = cur.fetchone()
    except:
        db.rollback()

    try:
        fech_gst=cur1.execute(querry_gest)
        db.commit()
        registro1 = cur1.fetchone()
        while (registro1 != None):
            fech_gst=registro1[0]
            registro1 = cur1.fetchone()
    except:
        db.rollback()

    try:
        contar=cur2.execute(contar_gestion)
        db.commit()
        registro2 = cur2.fetchone()
        while (registro2 != None):
            cont=registro2[0]
            registro2 = cur2.fetchone()
    except:
        db.rollback()


    hoja2.cell(row=i+1,column=1).value=cedula
    hoja2.cell(row=i+1,column=2).value=seg
    hoja2.cell(row=i+1,column=3).value=aplData
    hoja2.cell(row=i+1,column=6).value=fec_ivr
    hoja2.cell(row=i+1,column=7).value=fech_gst
    hoja2.cell(row=i+1,column=8).value=nec_gst
    hoja2.cell(row=i+1,column=9).value=cont
    hoja2.cell(row=i+1,column=12).value=valor
    hoja2.cell(row=i+1,column=13).value=val_hon
    a=a+1
    print(a)


db.close()
doc.save("PAGOS.xlsx")
